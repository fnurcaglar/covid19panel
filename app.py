# -*- coding: utf-8 -*-
"""
Created on Thu Aug 26 17:19:03 2020
Edited on Thu Sep 3 01:32:13 2020

@author: Fatih Erdem
"""

import subprocess
import sys

try:
    from flask import Flask, render_template, request, flash, url_for, redirect
except ImportError:
    subprocess.call([sys.executable, "-m", "pip", "install", 'flask'])
finally:
    from flask import Flask, render_template, request, flash, url_for, redirect

try:
    from flask_socketio import SocketIO
except ImportError:
    subprocess.call([sys.executable, "-m", "pip", "install", 'flask_socketio'])
finally:
    from flask_socketio import SocketIO

try:
    from flask_sqlalchemy import SQLAlchemy
except ImportError:
    subprocess.call([sys.executable, "-m", "pip", "install", 'flask_sqlalchemy'])
finally:
    from flask_sqlalchemy import SQLAlchemy

try:
    import serial
except ImportError:
    subprocess.call([sys.executable, "-m", "pip", "install", 'serial'])
finally:
    import serial

try:
    from glob import glob
except ImportError:
    subprocess.call([sys.executable, "-m", "pip", "install", 'glob'])
finally:
    from glob import glob

try:
    import threading
except ImportError:
    subprocess.call([sys.executable, "-m", "pip", "install", 'threading'])
finally:
    import threading

try:
    import locale
except ImportError:
    subprocess.call([sys.executable, "-m", "pip", "install", 'locale'])
finally:
    import locale

try:
    import io
except ImportError:
    subprocess.call([sys.executable, "-m", "pip", "install", 'io'])
finally:
    import io

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    subprocess.call([sys.executable, "-m", "pip", "install", 'openpyxl'])
finally:
    from openpyxl import Workbook, load_workbook

import shutil
import serial.tools.list_ports
import math
from database import *
import random
from datetime import datetime
import os
from avoserial import *
import time

app = Flask(__name__)
app.config["SECRET_KEY"] = "secret!"
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///avokadio.db'
locale.setlocale(locale.LC_ALL, 'tr_TR')
sio = SocketIO(app, logger = False)
baudrate = 76800
tema = "light"
samplePerSecond = 2
measureTime = 100
plotStatus = 0
percentage = 0

db = SQLAlchemy(app)
print_format = ""

class users(db.Model):
	id 			= db.Column(db.Integer, unique=True, primary_key=True)
	userno 		= db.Column(db.String(10)) #TODO: change to integer
	name 		= db.Column(db.String(10))
	phone 		= db.Column(db.Integer)
	gender 		= db.Column(db.String(10))
	height 		= db.Column(db.Integer)
	weight		= db.Column(db.Integer)
	leftpnif	= db.Column(db.Integer) 
	rightpnif	= db.Column(db.Integer) 
	globalpnif	= db.Column(db.Integer)
	birthyear	= db.Column(db.Integer)
	note		= db.Column(db.Integer)
	#date		= db.Column(db.DateTime,default=datetime.utcnow)


config = {
		"host":"0.0.0.0",
		"port":5000,
		# "ip":"192.168.1.104:5000"
		"ip":"127.0.0.1:5000"
}

linkler = {
  "olcum": "olcum",
#   "ayarlar": "ayarlar",
  "yenikayit": "yenikayit",
  "kullanicilar": ""
}

def is_integer(n):
    try:
        float(n)
    except ValueError:
        return False
    else:
        return float(n).is_integer()

def list_users():
	global numberofusers
	global biggestid

	users_local = users.query.all()
	users_list = []

	for userindb in users_local:
		try:
			user = {
				"id": userindb.id,
				"userno": userindb.userno,
				"name": userindb.name,
				"phone": userindb.phone,
				"gender": userindb.gender,
				"height": userindb.height,
				"weight": userindb.weight,
				"leftpnif": userindb.leftpnif,
				"rightpnif": userindb.rightpnif,
				"birthyear": userindb.birthyear,
				"note": userindb.note,
			}
			user_list.append(user)
		except:
			pass
			# print("User doesn't exists at ID = {}".format(u+1))

	numberofusers = len(user_list)
	
	print("Biggest number is {}, yet there are {} users in the database.".format(biggestid, numberofusers))

	return user_list


#user_list = list_users()

def saveToExcel(sheet,data):
	veri = data.split(',')
	satir_no = int(veri[1])
	sheet.cell(row=satir_no+2, column=1).value = veri[0]
	for i in range(1,9):
		sheet.cell(row=satir_no+2, column=i+1).value = float(veri[i])

# @app.route("/")
# def anasayfa():
# 	return render_template("index.html", ip=config["ip"] ,config=config, tema=tema)

@app.route("/ayarlar")
def ayarlar():
	return render_template("settings.html", ip=config["ip"] ,config=config, tema=tema, url=linkler["ayarlar"])

@app.route("/olcum", methods=['POST', 'GET'])
def olcum():
	if request.method == 'POST':
		now = datetime.now().strftime("%y%m%d-%H%M%S")
		print(now)

		no = request.form.get('userno')
		userno = no.split()[0]
		
		user = users.query.filter_by(userno=userno).first()
		return render_template("plot.html", ip=config["ip"] ,config=config, tema=tema, url=linkler["olcum"], user=user, users=users.query.all(),  sure=measureTime, plotStatus=plotStatus, percentage=percentage, alerts=[0,0])
	else:
		return render_template("plot.html", ip=config["ip"] ,config=config, tema=tema, url=linkler["olcum"],users=users.query.all() ,  sure=measureTime, plotStatus=plotStatus, percentage=percentage, alerts=[0,0])

@app.route("/")
def kullanicilar():
	return render_template("users.html", ip=config["ip"] ,config=config,users=users.query.all(), tema=tema, url=linkler["kullanicilar"])

@app.route("/kullanicilar/<userno>")
def kullanici(userno):
	addr = "static//recs//"+userno+"//excel"
	fileaddresses = glob(addr+"//*")
	fileaddresses.reverse()
	filenames = []
	for file in fileaddresses:
		file = file.replace(addr,"")
		print(file)
		file = file[1:(len(file)-5)]
		print(file)
		name = datetime.strptime(file, '%y%m%d-%H%M%S')	
		file = name.strftime('%y%m%d-%H%M%S')
		filenames.append(file)
	numberofrecords = len(fileaddresses)
	user = users.query.filter_by(userno=userno).first()
	return render_template("user.html", ip=config["ip"] ,config=config, tema=tema, url=linkler["olcum"], user=user, names=filenames, addresses=fileaddresses, number=numberofrecords)

@app.route("/yenikayit")
def kayit():
	return render_template("yenikayit.html", ip=config["ip"] ,config=config, tema=tema, url=linkler["yenikayit"])

@app.route('/addrec',methods = ['POST', 'GET'])
def addrec():
	global user_list
	if request.method == 'POST':
		userno = request.form['userno']
		name = request.form['name']
		phone = request.form['phone']
		gender = request.form['gender']
		height = request.form['height']
		weight = request.form['weight']
		leftpnif = request.form['leftpnif']
		rightpnif = request.form['rightpnif']
		globalpnif = request.form['globalpnif']
		birthyear = request.form['birthyear']
		note = request.form['note']

		new_user = users(userno=userno, name=name, phone=phone, gender=gender, height=height, weight=int(float(weight)*10), leftpnif=leftpnif, rightpnif=rightpnif, globalpnif=globalpnif, birthyear=birthyear, note=note)
		db.session.add(new_user)
		db.session.commit()

		#user_list = list_users()

		new = name

	return render_template("yenikayit.html", ip=config["ip"] ,config=config, tema=tema, new=new)

@app.route('/olcum/<userno>')
def newrec(userno):

	now = datetime.now().strftime("%y%m%d-%H%M%S")
	print(now)

	user = users.query.filter_by(userno=userno).first()
	return render_template("plot.html", ip=config["ip"] ,config=config, tema=tema, url=linkler["olcum"], user=user, users=users.query.all(), sure=measureTime, plotStatus=plotStatus, percentage=percentage, alerts=[0,0])

@sio.on("deleteUser")
def deleteUser(userno):
	#global user_list

	db.session.query(users).filter(users.userno==userno).delete()
	db.session.commit()

	#user_list = list_users()

	return render_template("users.html", ip=config["ip"] ,config=config, tema=tema, url=linkler["kullanicilar"], users= users.query.all())

@sio.on("startMeasure")
def startMeasure(userno):
	global plotStatus
	global percentage
	global measureTime
	global samplePerSecond

	print(userno, measureTime)

	now = datetime.now().strftime("%y%m%d-%H%M%S")
	print(now)

	totalSample = int(measureTime*samplePerSecond)
	print(totalSample)

	try:
		os.mkdir("static//recs//"+userno+"//csv//")
		os.mkdir("static//recs//"+userno+"//excel//")
	except:
		print("FOLDER EXISTS")

	#try:
	print("COM21")
	initAvoSerial("COM21", 57600)
	sio.emit("startMeasure", [1, 0])
	f = open("static//recs//"+userno+"//csv//"+now+".csv", "a")
	newPath = shutil.copy('user_records.xlsx', "static//recs//"+userno+"//excel//"+now+".xlsx")
	workbook = load_workbook(filename="static//recs//"+userno+"//excel//"+now+".xlsx")
	print (workbook.sheetnames)
	sheet = workbook["records"]
	f.write("time,number,nh3,co,no2,c3h8,c4h10,ch4,h2,c2h5oh\n")
	startMeasureBy(totalSample)
	data = readData()
	while (data != "STARTED"):
		startMeasureBy(totalSample)
		data = readData()
	for d in range(totalSample):
		data = readData()
		id_no = data.split(',')[1] 
		if is_integer(id_no): #is_integer fonksiyonu string içinde int varsa True dönüyor.
			if (int(id_no) == d):
				print(data)
				f.write(data+'\n')
				saveToExcel(sheet,data)
				sio.emit("startMeasure", [1, int(100*(d+1)/(totalSample))])
			else:
				print("DATA READ ERROR! TRY AGAIN")
	f.close()
	workbook.save(filename="static//recs//"+userno+"//excel//"+now+".xlsx")
	sio.emit("startMeasure", [0, 100])
	#TODO: basari ile tamamlandi diye arayuze ekleyecegiz.
	closeAvoSerial()
	# except:
	# 	print("CANNOT CONNECTED")

	sio.emit("alert", [1,1])
	sio.emit("startMeasure", [0, 0])

@sio.on("changeDuration")
def changeDuration(data):
	global measureTime
	measureTime = int(data)
	print(measureTime)

@sio.on("connect")
def connect():
	print("Connected")

@sio.on("disconnect")
def disconnect():
	print("Disconnected")

@sio.on("changeTheme")
def changeTheme():
	global tema
	if tema == "dark":
		tema = "light"
	elif tema == "light":
		tema = "dark"
	sio.emit("changeTheme", tema)

if __name__ == "__main__":
	app.debug = True
	print("\n\n - - - - - - - - - - - - - - - - - - - - - - - - - - -\nAvokadio\n - - - - - - - - - - - - - - - - - - - - - - - - - - -\n\n")
	# print("                           dP                      dP oo          \n                           88                      88             \n.d8888b. dP   .dP .d8888b. 88  .dP  .d8888b. .d888b88 dP .d8888b. \n88'  `88 88   d8' 88'  `88 88888    88'  `88 88'  `88 88 88'  `88 \n88.  .88 88 .88'  88.  .88 88  `8b. 88.  .88 88.  .88 88 88.  .88 \n`88888P8 8888P'   `88888P' dP   `YP `88888P8 `88888P8 dP `88888P' \n")
	sio.run(app, host=config["host"], port=config["port"])